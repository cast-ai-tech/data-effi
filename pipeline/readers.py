"""File readers: bytes in, (headers, rows) out.

Deliberately dumb. No business logic, no type coercion beyond what the file
format itself already decided (a spreadsheet cell that is a real date stays a
date). Everything else is normalize.py's job.

THREE FORMATS, AND ONE OF THEM LIES ABOUT ITSELF. Effi exports its wallet report
as `Reporte de movimientos ... .xls`, and that file is not Excel at all - it is
an HTML `<table>`. Trusting the extension gets you a corrupt-file error on a
perfectly good report, so every file is sniffed by content first.
"""

from __future__ import annotations

import csv
import html
import io
import re
from collections.abc import Iterator
from typing import Any

SUPPORTED_EXTENSIONS = (".csv", ".xlsx", ".xlsm", ".xls", ".txt", ".tsv", ".html", ".htm")

_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
_MAX_PREVIEW_BYTES = 64 * 1024

# Magic bytes. ZIP for OOXML, the OLE2 header for genuine legacy .xls.
_ZIP_MAGIC = b"PK\x03\x04"
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class UnsupportedFileError(ValueError):
    """The file is not something Norte knows how to read."""


class EmptyFileError(ValueError):
    """The file parsed fine but carries no usable rows."""


def sniff_format(payload: bytes, filename: str) -> str:
    """Decide how to read a file from its CONTENT, falling back to its name.

    Returns one of: 'xlsx', 'html', 'csv', 'xls_binary'.
    """
    head = payload[:2048]

    if head.startswith(_ZIP_MAGIC):
        return "xlsx"
    if head.startswith(_OLE2_MAGIC):
        return "xls_binary"

    # An HTML table, however the exporter chose to name the file.
    lowered = head.lstrip()[:512].lower()
    if lowered.startswith((b"<", b"\xef\xbb\xbf<")) or b"<table" in head.lower():
        return "html"

    lowered_name = filename.lower()
    if lowered_name.endswith((".xlsx", ".xlsm")):
        return "xlsx"
    if lowered_name.endswith((".html", ".htm")):
        return "html"
    return "csv"


def read_tabular(payload: bytes, filename: str) -> tuple[list[str], list[list[Any]]]:
    """Read a spreadsheet, an HTML table or a delimited file into headers + rows."""
    fmt = sniff_format(payload, filename)

    if fmt == "xlsx":
        headers, rows = _read_xlsx(payload)
    elif fmt == "html":
        headers, rows = _read_html_table(payload)
    elif fmt == "xls_binary":
        raise UnsupportedFileError(
            f"'{filename}' es un Excel binario antiguo (.xls real). "
            "Ábrelo en Excel y guárdalo como .xlsx, o expórtalo de nuevo."
        )
    else:
        headers, rows = _read_csv(payload)

    if not headers:
        raise EmptyFileError(f"{filename}: no se encontró una fila de encabezados")
    if not rows:
        raise EmptyFileError(f"{filename}: el archivo no tiene filas de datos")
    return headers, rows


def _decode(payload: bytes) -> str:
    last_error: Exception | None = None
    for encoding in _ENCODINGS:
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise UnsupportedFileError(f"No se pudo decodificar el archivo: {last_error}")


def _sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        counts = {sep: sample.count(sep) for sep in (";", ",", "\t", "|")}
        return max(counts, key=counts.get) if any(counts.values()) else ","


def _read_csv(payload: bytes) -> tuple[list[str], list[list[Any]]]:
    text = _decode(payload)
    delimiter = _sniff_delimiter(text[:_MAX_PREVIEW_BYTES])
    reader = csv.reader(io.StringIO(text, newline=""), delimiter=delimiter)

    headers: list[str] = []
    rows: list[list[Any]] = []
    for raw_row in reader:
        if not headers:
            if _is_blank(raw_row):
                continue
            headers = [str(cell).strip() for cell in raw_row]
            continue
        if _is_blank(raw_row):
            continue
        rows.append(list(raw_row))
    return headers, rows


def _read_xlsx(payload: bytes) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:      # pragma: no cover - dependency is declared
        raise UnsupportedFileError(
            "Falta la dependencia openpyxl para leer archivos Excel"
        ) from exc

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        headers: list[str] = []
        rows: list[list[Any]] = []
        for raw_row in sheet.iter_rows(values_only=True):
            values = list(raw_row)
            if not headers:
                if _is_blank(values):
                    continue
                headers = [("" if cell is None else str(cell).strip()) for cell in values]
                continue
            if _is_blank(values):
                continue
            rows.append(values)
        return headers, rows
    finally:
        workbook.close()


# -----------------------------------------------------------------------------
# HTML table reader
# -----------------------------------------------------------------------------

_ROW_RE = re.compile(r"<tr[^>]*>(.*?)</tr>", re.IGNORECASE | re.DOTALL)
_CELL_RE = re.compile(r"<t[dh][^>]*>(.*?)</t[dh]>", re.IGNORECASE | re.DOTALL)
_HEADER_CELL_RE = re.compile(r"<th[^>]*>(.*?)</th>", re.IGNORECASE | re.DOTALL)
_TAG_RE = re.compile(r"<[^>]+>")
_WS_RE = re.compile(r"\s+")


def _clean_cell(fragment: str) -> str:
    """Strip tags and entities, collapse whitespace.

    Effi embeds literal `<br>` inside note fields, so a single cell can span
    several lines. Collapsing them keeps the value one line without losing it.
    """
    text = _TAG_RE.sub(" ", fragment)
    text = html.unescape(text)
    return _WS_RE.sub(" ", text).strip()


def _read_html_table(payload: bytes) -> tuple[list[str], list[list[Any]]]:
    """Read the first HTML table in the document.

    Regex rather than a parser on purpose: these exports are machine-generated,
    flat, and up to a few hundred megabytes. A DOM parser would build an object
    per cell for a file that is only ever read once, top to bottom.
    """
    text = _decode(payload)

    raw_rows = _ROW_RE.findall(text)
    if not raw_rows:
        raise EmptyFileError("El archivo parece HTML pero no contiene una tabla")

    headers: list[str] = []
    rows: list[list[Any]] = []

    for raw_row in raw_rows:
        header_cells = _HEADER_CELL_RE.findall(raw_row)
        if not headers and header_cells:
            headers = [_clean_cell(cell) for cell in header_cells]
            continue

        cells = [_clean_cell(cell) for cell in _CELL_RE.findall(raw_row)]
        if not cells or all(cell == "" for cell in cells):
            continue

        # A table with no <th> at all: take the first row as the header.
        if not headers:
            headers = cells
            continue

        rows.append(cells)

    return headers, rows


def _is_blank(row: list[Any]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def iter_records(
    headers: list[str], rows: list[list[Any]], header_map: dict[int, str]
) -> Iterator[tuple[int, dict[str, Any], dict[str, Any]]]:
    """Yield (row_number, mapped_fields, raw_payload) for each data row.

    row_number is 1-based over data rows (the header is row 0), which is what the
    error report shows the user. raw_payload keeps the whole original row so it
    can be stored in raw.source_row for audit.
    """
    for index, row in enumerate(rows, start=1):
        mapped: dict[str, Any] = {}
        raw: dict[str, Any] = {}
        for position, value in enumerate(row):
            header = headers[position] if position < len(headers) else f"col_{position}"
            raw[header] = None if value is None else str(value)
            field_name = header_map.get(position)
            if field_name is not None:
                mapped[field_name] = value
        yield index, mapped, raw
